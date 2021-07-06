#
import mysql as mysql
import openpyxl

path = "students.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=2, column=2)
print(cell_obj.value)

#
# dbse = mydb.cursor()
# dbse.execute("CREATE DATABASE Students_Management_System1")
print("Students_Management_System Already Created")

#
dbse = mydb.cursor()
dbse.execute("SHOW DATABASES")
for entry in dbse:
    print(entry)

#
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="students_management_system1"
)
dbse = mydb.cursor()
dbse.execute("CREATE TABLE studentdata(rno VARCHAR(10),name VARCHAR(255),department VAR)CHAR(255) ,contact_no "
             "VARCHAR(255),sem1_marks VARCHAR(255),sem2_marks VARCHAR(255),average VARCHAR(255)")

#
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="students_management_system1"
)
dbse = mydb.cursor()
dbse.execute("SHOW TABLES")
for value in dbse:
    print(value)

#
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="students_management_system"
)
cur = mydb.cursor()
cur.execute('SELECT * FROM studentdata')
for row in cur:
    print(row)

#
import xlrd

xl_sheet = xlrd.open_workbook("students.xlsx")
xl_sheet
sheet_name = xl_sheet.sheet_names()
sheet_name

#
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="students_management_system"
)
cur = mydb.cursor()
for s in range(0, 1):
    sheet = xl_sheet.sheet_by_index(s)
    sql = "INSERT INTO studentdata(rno,name,department,contact_no,sem1_marks,sem2_marks,average) VALUES(%s,%s,%s,%s,%s,%s,%s)"
for r in range(1, sheet.nrows):
    rno = sheet.cell(r, 0).value
    name = sheet.cell(r, 1).value
    department = sheet.cell(r, 2).value
    contact_no = sheet.cell(r, 3).value
    sem1_marks = sheet.cell(r, 4).value
    sem2_marks = sheet.cell(r, 5).value
    average = sheet.cell(r, 6).value
    values = (rno, name, department, contact_no, sem1_marks, sem2_marks, average)
cur.execute(sql, values)
mydb.commit()
mycursor = mydb.cursor()
mycursor.execute("SELECT * FROM studentdata")
myresult = mycursor.fetchall()
for x in myresult:
    print(x)
mydb.commit()
mydb.close()
